"""
Converts the ranked CSV output into a formatted XLSX deliverable.

Usage:
    python export_xlsx.py --csv submission.csv --out ranked_candidates.xlsx

This exists so the XLSX deliverable (required by the Hack2Skill submission
portal) is reproducible from the repo itself, not a one-off artifact
produced outside the codebase.
"""

import argparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_NAVY = "1A2B4C"


def export_xlsx(csv_path: str, out_path: str) -> None:
    df = pd.read_csv(csv_path)
    df.to_excel(out_path, index=False, sheet_name="Ranked Candidates")

    wb = load_workbook(out_path)
    ws = wb["Ranked Candidates"]

    header_fill = PatternFill("solid", start_color=HEADER_NAVY, end_color=HEADER_NAVY)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {"A": 16, "B": 8, "C": 10, "D": 110}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column_letter == "D"))

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 30

    wb.save(out_path)
    print(f"Wrote {out_path} ({ws.max_row - 1} candidates)")


def main():
    p = argparse.ArgumentParser(description="Export the ranked CSV to a formatted XLSX")
    p.add_argument("--csv", default="submission.csv", help="Path to the ranked CSV (output of main.py)")
    p.add_argument("--out", default="ranked_candidates.xlsx", help="Output XLSX path")
    args = p.parse_args()
    export_xlsx(args.csv, args.out)


if __name__ == "__main__":
    main()
